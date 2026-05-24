"""
Expanded document loader service.
Supports PDF, TXT, DOCX, CSV, Excel, Markdown, and web URLs.
"""

import os
import tempfile
from pathlib import Path
from typing import List

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import PyPDFLoader, TextLoader


SUPPORTED_FORMATS = {
    ".pdf": "PDF Document",
    ".txt": "Text File",
    ".docx": "Word Document",
    ".csv": "CSV Spreadsheet",
    ".xlsx": "Excel Spreadsheet",
    ".xls": "Excel Spreadsheet (Legacy)",
    ".md": "Markdown File",
}


def get_supported_formats() -> dict:
    """Return dict of supported file extensions and their descriptions."""
    return SUPPORTED_FORMATS.copy()


def load_document(file_path: str, file_type: str) -> list:
    """
    Load a document from a file path based on its type.
    Returns a list of LangChain Document objects.
    """
    suffix = file_type.lower()

    if suffix == ".pdf":
        try:
            from langchain_community.document_loaders import UnstructuredPDFLoader
            loader = UnstructuredPDFLoader(file_path, mode="elements", strategy="hi_res")
            return loader.load()
        except Exception as e:
            print(f"Unstructured fallback due to: {e}")
            # Fallback
            loader = PyPDFLoader(file_path)
            return loader.load()

    elif suffix == ".txt":
        loader = TextLoader(file_path, encoding="utf-8")
        return loader.load()

    elif suffix == ".docx":
        try:
            from langchain_community.document_loaders import Docx2txtLoader
            loader = Docx2txtLoader(file_path)
            return loader.load()
        except ImportError:
            raise ImportError("Install docx2txt: pip install docx2txt")

    elif suffix == ".csv":
        try:
            from langchain_community.document_loaders import CSVLoader
            loader = CSVLoader(file_path, encoding="utf-8")
            return loader.load()
        except ImportError:
            raise ImportError("CSVLoader not available in langchain-community")

    elif suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            from langchain.schema import Document

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            docs = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    rows.append(row_str)
                content = "\n".join(rows)
                if content.strip():
                    docs.append(Document(
                        page_content=content,
                        metadata={"source": file_path, "sheet": sheet_name}
                    ))
            wb.close()
            return docs
        except ImportError:
            raise ImportError("Install openpyxl: pip install openpyxl")

    elif suffix == ".md":
        # Simple markdown loading as text
        loader = TextLoader(file_path, encoding="utf-8")
        return loader.load()

    else:
        raise ValueError(f"Unsupported file format: {suffix}")


def load_url(url: str, depth: int = 1) -> list:
    """Load content from a web URL, optionally crawling recursively."""
    try:
        if depth <= 1:
            from langchain_community.document_loaders import WebBaseLoader
            loader = WebBaseLoader(url)
            return loader.load()
        else:
            from langchain_community.document_loaders import RecursiveUrlLoader
            from bs4 import BeautifulSoup
            
            def extractor(x):
                return BeautifulSoup(x, "html.parser").get_text()
                
            loader = RecursiveUrlLoader(url, max_depth=depth, extractor=extractor)
            return loader.load()
    except ImportError:
        raise ImportError("Install beautifulsoup4: pip install beautifulsoup4")


def split_documents(docs: list, chunk_size: int = 1000, chunk_overlap: int = 200) -> list:
    """Split documents into chunks for embedding."""
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ".", " ", ""],
    )
    return text_splitter.split_documents(docs)
